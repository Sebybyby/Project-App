import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os


class ExcelFileCreator:
    def __init__(self, filename):
        super().__init__()
        self.filename = filename
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Rapport Excel"
        
    def create_file(self):
        # En-têtes des colonnes
        headers = ["Timestamp", "Group", "Index", "Row", "Col", "Trans", "Color"]
        self.ws.append(headers)
        self.wb.save(self.filename)
        print(f"Fichier Excel '{self.filename}' créé avec succès.")
    
    def write_data(self, data):
        self.ws.append(data)
        self.wb.save(self.filename)
        print(f"Données écrites dans '{self.filename}': {data}")
    
    def color_cell(self, row, col, color):
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell = self.ws.cell(row=row, column=col)
        cell.fill = fill
        self.wb.save(self.filename)
        print(f"Cellule ({row}, {col}) colorée en {color} dans '{self.filename}'.")
    
    def close(self):
        self.wb.save(self.filename)
        print(f"Fichier Excel '{self.filename}' sauvegardé et fermé.")
#  Exemple d'utilisation
# excel_creator = ExcelFileCreator("lane5000_results.xlsx")
# excel_creator.create_file()
# excel_creator.write_data(["2024-06-01 12:00:00", "A", 1, 0, 2, "True", "Green"])
# excel_creator.color_cell(2, 7, "00FF00")  # Colorie la cellule G2 en vert
# excel_creator.close()

    